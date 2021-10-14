import ipaddress
import openpyxl
import dns.resolver


if __name__ == '__main__':
    # Alinan input excel dosyasından ipleri okur ve
    # DNS A Kayitlarini Cozerek Ayrı bir dosyaya yazar.
    
    # Input Dosyası
    wb_obj = openpyxl.load_workbook('input.xlsx')
    sheet_obj = wb_obj.active
    
    # Output Dosyası
    wb = openpyxl.Workbook()
    sheet = wb.active

    for i in range(1, sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i, column=1)
        dnsValue = cell_obj.value
        dnsAResult = ""
        try:
            dnsAResultDict = dns.resolver.query(dnsValue, 'A')
            for val in dnsAResultDict:
                dnsAResult = val.to_text()
        except:
            dnsAResult = "hata: "+dnsValue

        sheet.cell(row=i, column=1).value = str(dnsValue)
        sheet.cell(row=i, column=2).value = str(dnsAResult)
        print(str(i)+"/"+str(sheet_obj.max_row))

    wb.save("result.xlsx")
