import ipaddress
import openpyxl

if __name__ == '__main__':
    # Program input olarak iki dosya alir.
    # Input 1: sahipsizSistemler.xlsx
    # Input 2: sahipler.xlsx
    # Birinci input'daki tum ip ler sırayla okunur. O ip digerki dosyada ilk sutundeki subnette mi sorgulanir.
    # eger sorgulama dogruysa result.xlsx dosyasına hepsi yazilir.


    yeniIP = ""
    wb_obj = openpyxl.load_workbook('sahipsizSistemler.xlsx')
    sheet_obj = wb_obj.active
    wb_obj2 = openpyxl.load_workbook('sahipler.xlsx')
    sheet_obj2 = wb_obj2.active

    #Result File
    wb = openpyxl.Workbook()
    sheet = wb.active

    m = 1

    # i: Birinci inputtaki her bir satırı ifade eder.
    for i in range(1, sheet_obj.max_row):
        cell_obj = sheet_obj.cell(row=i, column=1)
        # eger nokta yoksa excel den kaynaklanir, bunlari duzeltmek icin ip yapılır.
        if "." not in str(cell_obj.value):
            #print(cell_obj.value)
            ipString = str(cell_obj.value)
            yeniIP = ipString[0:2]+"."+ipString[2:5]+"."+ipString[5:8]+"."+ipString[8:11]
            #print(yeniIP)
        else:
            yeniIP = str(cell_obj.value)

        # input 1 deki her satir icin ikinci exceldeki her satır sorgulanir.
        for j in range(2,sheet_obj2.max_row):
            cell_obj2 = sheet_obj2.cell(row=j,column=1)

            # a: İkinci inouttaki network
            # b: Birinci inputtaki ip adresi /32
            a = ipaddress.ip_network(str(cell_obj2.value).strip())
            b = ipaddress.ip_network(yeniIP)

            # Eger ip ilgili networkteyse
            if b.subnet_of(a) == True:
                sheet.cell(row=m, column=1).value = str(b)
                for k in range(1,sheet_obj2.max_column):
                    c3 = sheet_obj2.cell(row=j,column=k).value
                    sheet.cell(row=m, column=k+1).value = str(c3)
                    #print(str(b)+":"+str(c3))
                m = m + 1

    wb.save("result.xlsx")
